# app.py — GST Reconciliation Tool Enterprise v4.0
# ══════════════════════════════════════════════════════
# CLOUD LOGIN GATE — Each client has their own unique key
# ══════════════════════════════════════════════════════
import streamlit as _st
import platform as _platform
import hashlib as _hashlib

def _cloud_login_gate():
    """
    On cloud (Linux): Each client must enter their unique activation key.
    - Every client has a DIFFERENT key from key_hashes.py
    - If a client shares their key, you can revoke it by removing that
      hash from key_hashes.py and pushing to GitHub — app updates in 2 min.
    - Key is stored in session — user must re-enter only when browser is closed.
    On local EXE (Windows): This gate is completely skipped.
    """
    if _platform.system() != "Linux":
        return  # Local EXE — skip cloud gate entirely

    from modules.key_hashes import VALID_KEY_HASHES

    # Already logged in this session — let them through
    if _st.session_state.get("cloud_authenticated"):
        return

    # ── Show Login Screen ─────────────────────────────────────────────────────
    _st.set_page_config(
        page_title="GST Tool — Login",
        page_icon="🔐",
        layout="centered"
    )

    _st.markdown("""
    <style>
    .stApp {
        background: linear-gradient(135deg, #0f2044 0%, #1a3a6e 50%, #0f52ba 100%);
    }
    .login-card {
        background: white;
        border-radius: 24px;
        padding: 48px 52px;
        max-width: 460px;
        margin: 50px auto 0 auto;
        box-shadow: 0 24px 64px rgba(0,0,0,0.35);
        text-align: center;
    }
    .login-title { color: #0f2044; font-size: 26px; font-weight: 700; margin-bottom: 4px; }
    .login-sub   { color: #666; font-size: 14px; margin-bottom: 28px; }
    .key-note    { color: #888; font-size: 12px; margin-top: 16px; }
    </style>
    """, unsafe_allow_html=True)

    _st.markdown("""
    <div class="login-card">
        <div style="font-size:52px;">🔐</div>
        <div class="login-title">GST Reconciliation Tool</div>
        <div class="login-sub">Enterprise v4.0 — Please enter your access key</div>
    </div>
    """, unsafe_allow_html=True)

    _st.markdown("### Enter Your Unique Access Key")
    _st.caption("Each client has their own personal key. Contact the seller if you don't have one.")

    _key_input = _st.text_input(
        "Access Key",
        placeholder="XXXX-XXXX-XXXX-XXXX",
        max_chars=50,
        label_visibility="collapsed"
    )

    col1, col2, col3 = _st.columns([1, 2, 1])
    with col2:
        login_btn = _st.button("🔓 Access App", type="primary", use_container_width=True)

    if login_btn:
        if not _key_input.strip():
            _st.error("⚠️ Please enter your access key.")
        else:
            # Hash the entered key same way as license_manager
            _entered_hash = _hashlib.sha256(
                _key_input.strip().upper().encode()
            ).hexdigest()

            if _entered_hash in VALID_KEY_HASHES:
                # ✅ Valid key — store in session
                _st.session_state["cloud_authenticated"] = True
                _st.session_state["cloud_key_hash"] = _entered_hash
                _st.success("✅ Access granted! Loading app...")
                _st.rerun()
            else:
                _st.error("❌ Invalid key. Please check and try again, or contact the seller.")
                _st.info("💡 Each key is unique to one client. If you received this link from someone else, you need your own key.")

    _st.markdown("---")
    _st.markdown(
        "<p style='text-align:center; color:#aaa; font-size:12px;'>"
        "🔒 Your key is personal. Do not share it — shared keys can be revoked at any time."
        "</p>",
        unsafe_allow_html=True
    )
    _st.stop()  # ← Hard stop. Nothing below runs until key is entered.

# Run the gate FIRST before anything else
_cloud_login_gate()

# ══════════════════════════════════════════════════════
# LICENSE GATE — runs before any app logic
# ══════════════════════════════════════════════════════

# We need page config first, so set it here early if not already set
# (The real set_page_config is called again below — Streamlit ignores duplicates)
try:
    from modules.license_manager import get_license_status, activate_key, is_allowed_to_run

    _lic = get_license_status()

    # ── Blocked (wrong device) or expired — show gate and STOP ────────────────
    if _lic["status"] in ("expired_trial", "expired_key", "blocked"):
        _st.set_page_config(page_title="GST Tool — Activation Required", page_icon="🔐", layout="centered")
        _st.markdown("""
        <style>
        .stApp { background: linear-gradient(135deg, #1e3a5f 0%, #0f52ba 100%); }
        .act-card {
            background: white; border-radius: 20px; padding: 40px 50px;
            max-width: 500px; margin: 60px auto; box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            text-align: center;
        }
        </style>""", unsafe_allow_html=True)

        _st.markdown("""
        <div class="act-card">
            <h1>🔐 GST Reconciliation Tool</h1>
            <h3 style="color:#e74c3c;">Activation Required</h3>
        </div>""", unsafe_allow_html=True)

        _st.error(f"🚫 {_lic['message']}")
        _st.info(f"📟 Your Device ID: `{_lic['mac']}`")
        _st.markdown("---")
        _st.markdown("### Enter Your Activation Key")
        _st.caption("Purchase a key from the seller and enter it below.")

        _key_input = _st.text_input("Activation Key", placeholder="XXXX-XXXX-XXXX-XXXX", max_chars=19)
        if _st.button("🔓 Activate Software", type="primary", use_container_width=True):
            if _key_input:
                _result = activate_key(_key_input)
                if _result["success"]:
                    _st.success(_result["message"])
                    _st.balloons()
                    _st.rerun()
                else:
                    _st.error(_result["message"])
            else:
                _st.warning("Please enter your activation key first.")
        _st.stop()  # ← Hard stop. App code below does NOT run.

    # ── Trial or Active — show small banner and continue ──────────────────────
    elif _lic["status"] == "trial":
        _trial_days = _lic["days_left"]
        # Store banner info in session for display after page config
        import streamlit as _st2
        if "lic_banner" not in _st2.session_state:
            _st2.session_state["lic_banner"] = ("trial", _lic["message"])

    elif _lic["status"] == "active":
        if "lic_banner" not in _st.session_state:
            _st.session_state["lic_banner"] = ("active", _lic["message"])

except Exception as _lic_err:
    pass  # If license module fails, allow run (fail open for safety)
# ══════════════════════════════════════════════════════
# END LICENSE GATE
# ══════════════════════════════════════════════════════
# CHANGES FROM v3.2:
#   BUG FIXES:
#     1. Unique_ID assigned before process_dataset() — manual links stable across re-runs
#     2. failed_matches no longer double-counted (core_engine fix)
#     3. Tab 2 Suggestions filter fixed (was if/elif logic bug)
#     4. gst_scraper import wrapped in try/except — no crash if file missing
#     5. CDNR results now saved to DB and restored when loading from History
#     6. K4 numeric key now has 4-digit minimum + tolerance check (false-match prevention)
#   FEATURES ADDED:
#     7. GSTIN format validator in setup stage
#     8. Template download for standard column mapping
#     9. Per-vendor tolerance override table
#    10. Combined B2B + CDNR Net ITC banner on Dashboard
#    11. Data confidence panel (Books vs GSTR-2B row/value summary) after upload
#    12. Executive Summary in Excel report (v4 report_gen)
#    13. CDNR Suggestions sub-view showing GSTIN match status
#    14. ITC Impact section removed from CDNR tab
#    15. Audit log for manual actions
#    16. Streamlit-lottie removed — uses native progress bar

import streamlit as st
import pandas as pd
import numpy as np
import altair as alt
import time
import urllib.parse
import io
import zipfile
import os
import re

# --- CORE IMPORTS ---
from modules.constants import (REQUIRED_FIELDS, FIXED_BOOKS_MAPPING, FIXED_GST_MAPPING,
                               SOFTWARE_COLUMN_PROFILES)
from modules.data_utils     import (load_data_preview, find_best_match,
                                    extract_meta_from_readme, standardize_invoice_numbers)
from modules.core_engine    import run_reconciliation
from modules.report_gen     import generate_excel, generate_vendor_split_zip
from modules.utils          import show_processing_animation
from modules.email_tool     import get_vendors_with_issues, generate_email_draft, generate_whatsapp_message
from modules.pdf_gen        import create_vendor_pdf
from modules.db_handler     import (init_db, save_reconciliation, get_history_list,
                                    load_reconciliation, delete_reconciliation,
                                    save_cdnr_to_history, log_action, get_audit_log)
from modules.file_manager   import get_client_path, save_file_to_folder, open_folder

# --- PRE-PROCESSORS ---
from modules.pre_processor  import smart_read_b2ba, process_amendments

# --- CDNR ENGINE ---
from modules.cdnr_processor    import process_cdnr_reconciliation
from modules.cdnr_report_gen   import generate_cdnr_excel

# ==========================================
# PAGE CONFIG & CSS
# ==========================================
st.set_page_config(
    page_title="GST Reconciliation Tool",
    page_icon="🛡️",
    layout="wide",
    initial_sidebar_state="expanded"
)
init_db()

st.markdown("""
    <style>
    .stApp { background-color: #f4f6f9; }
    div[data-testid="stMetric"] {
        background-color: #ffffff;
        border: 1px solid #e0e0e0;
        padding: 20px;
        border-radius: 12px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.04);
        transition: all 0.3s ease;
    }
    div[data-testid="stMetric"]:hover {
        transform: translateY(-4px);
        box-shadow: 0 8px 15px rgba(0,0,0,0.1);
        border-color: #0f52ba;
    }
    div[data-testid="stMetricLabel"] { font-size: 14px; color: #6c757d; font-weight: 600; text-transform: uppercase; letter-spacing: 0.5px; }
    div[data-testid="stMetricValue"] { font-size: 26px; color: #2c3e50; font-weight: 800; }
    .stTabs [data-baseweb="tab-list"] { gap: 10px; margin-bottom: 20px; }
    .stTabs [data-baseweb="tab"] {
        height: 45px; background-color: #ffffff; border-radius: 8px;
        padding: 0 24px; font-weight: 600; color: #495057;
        border: 1px solid #dee2e6; box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }
    .stTabs [data-baseweb="tab"][aria-selected="true"] {
        background-color: #0f52ba; color: #ffffff; border-color: #0f52ba;
    }
    div.stButton > button:first-child { border-radius: 8px; font-weight: 600; padding: 0.5rem 1rem; }
    div.stButton > button[kind="primary"] { background-color: #0f52ba; border: none; box-shadow: 0 4px 6px rgba(15, 82, 186, 0.3); }
    .streamlit-expanderHeader { background-color: white; border-radius: 8px; border: 1px solid #f0f0f0; }
    .main-header { font-family: 'Helvetica Neue', sans-serif; color: #1a1a1a; font-weight: 700; margin-bottom: 0px; }
    .sub-header { font-family: 'Helvetica Neue', sans-serif; color: #666; font-size: 16px; margin-bottom: 20px; }
    .confidence-box { background: #e8f5e9; border-left: 4px solid #4caf50; padding: 12px 16px; border-radius: 8px; margin: 8px 0; }
    .confidence-title { font-weight: 700; color: #2e7d32; font-size: 15px; margin-bottom: 6px; }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# SESSION STATE INIT
# ==========================================
defaults = {
    'app_stage':           'setup',
    'manual_matches':      [],
    'current_client_path': None,
    'gst_scraper':         None,
    'captcha_img':         None,
    'target_gstin':        None,
    'cdnr_result':         None,
    'cdnr_summary':        None,
    'current_recon_id':    None,
    'vendor_tolerances':   {},
    'data_summary_books':  None,
    'data_summary_gst':    None,
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ==========================================
# HELPERS
# ==========================================
def save_callback(folder_path, file_name, file_data):
    if folder_path:
        save_file_to_folder(folder_path, file_name, file_data)

def validate_gstin(gstin: str) -> bool:
    pattern = r'^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$'
    return bool(re.match(pattern, str(gstin).strip().upper()))

def make_data_summary(df, label):
    """Build quick stats dict from a raw DataFrame."""
    try:
        gstin_col = next((c for c in df.columns if 'gstin' in c.lower()), None)
        tax_col   = next((c for c in df.columns if 'taxable' in c.lower()), None)
        inv_col   = next((c for c in df.columns if 'invoice' in c.lower() and 'number' in c.lower()), None)
        n_rows    = len(df)
        n_gstin   = df[gstin_col].nunique() if gstin_col else '—'
        total_tax = pd.to_numeric(df[tax_col].astype(str).str.replace(',',''), errors='coerce').sum() if tax_col else 0
        n_inv     = df[inv_col].nunique() if inv_col else n_rows
        return {'label': label, 'rows': n_rows, 'invoices': n_inv, 'gstins': n_gstin, 'taxable': total_tax}
    except Exception:
        return {'label': label, 'rows': '?', 'invoices': '?', 'gstins': '?', 'taxable': 0}

def make_template_excel():
    """Generates a blank Purchase Register template Excel."""
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='xlsxwriter') as writer:
        headers = list(FIXED_BOOKS_MAPPING.values())
        headers = [h for h in headers if h != '<No Column / Blank>']
        sample  = {
            'GSTIN of Supplier': '24ABCDE1234F1Z5',
            'Invoice Number':     'INV-001',
            'Invoice date':       '01/04/2025',
            'Invoice Value':      '11800',
            'Taxable Value':      '10000',
            'Integrated Tax Paid': '1800',
            'Central Tax Paid':   '',
            'State/UT Tax Paid':  '',
            'Cess Paid':          '',
            'Place Of Supply':    '24-Gujarat',
            'Reverse Charge':     'N',
        }
        df_tmpl = pd.DataFrame([sample])
        df_tmpl.to_excel(writer, sheet_name='Purchase Register', index=False)
        wb  = writer.book
        ws  = writer.sheets['Purchase Register']
        hdr = wb.add_format({'bold': True, 'bg_color': '#4472C4', 'font_color': 'white', 'border': 1})
        for i, col in enumerate(df_tmpl.columns):
            ws.write(0, i, col, hdr)
            ws.set_column(i, i, 22)
    out.seek(0)
    return out.getvalue()

# ==========================================
# SIDEBAR — HISTORY
# ==========================================
with st.sidebar:
    st.markdown("### 🗂️ Reconciliation History")
    search_query = st.text_input("🔍 Search Client", placeholder="Type client name...").lower()
    history_df   = get_history_list()

    if not history_df.empty:
        if search_query:
            history_df = history_df[
                history_df['company_name'].str.lower().str.contains(search_query, na=False) |
                history_df['gstin'].str.lower().str.contains(search_query, na=False)
            ]
        if history_df.empty:
            st.warning("No matching records found.")
        else:
            st.markdown("---")
            for idx, row in history_df.iterrows():
                with st.container():
                    st.markdown(f"**{row['company_name']}**")
                    st.caption(f"📅 {row.get('fy','—')} | {row['period']} | {str(row['timestamp'])[:10]}")
                    c_open, c_del = st.columns([3, 1])
                    with c_open:
                        if st.button("📂 Open", key=f"hist_{row['id']}", use_container_width=True):
                            meta, df_loaded, df_cdnr, cdnr_summary = load_reconciliation(row['id'])
                            st.session_state['last_result']      = df_loaded
                            st.session_state['df_b_clean']       = df_loaded
                            st.session_state['df_g_clean']       = df_loaded
                            st.session_state['meta_gstin']       = meta['gstin']
                            st.session_state['meta_name']        = meta['company_name']
                            st.session_state['meta_fy']          = meta['fy']
                            st.session_state['meta_period']      = meta['period']
                            st.session_state.current_client_path = get_client_path(
                                meta['company_name'], meta['gstin'], meta['fy'], meta['period'])
                            st.session_state.current_recon_id    = row['id']
                            # Restore CDNR results from DB if available
                            st.session_state.cdnr_result  = df_cdnr
                            st.session_state.cdnr_summary = cdnr_summary
                            # Clear file bytes (not available for history loads)
                            st.session_state['file_books_bytes'] = None
                            st.session_state['file_gst_bytes']   = None
                            st.session_state.app_stage = 'results'
                            st.rerun()
                    with c_del:
                        if st.button("🗑️", key=f"del_{row['id']}", use_container_width=True, help="Delete permanently"):
                            delete_reconciliation(row['id'])
                            st.rerun()
                    st.markdown("""<hr style="margin:5px 0;border-color:#f0f0f0;">""", unsafe_allow_html=True)
    else:
        st.info("No saved history available.")

    # Audit log viewer
    if st.session_state.current_recon_id:
        with st.expander("📋 Audit Log", expanded=False):
            audit_df = get_audit_log(st.session_state.current_recon_id)
            if not audit_df.empty:
                for _, arow in audit_df.iterrows():
                    st.caption(f"**{arow['action_type']}** — {str(arow['timestamp'])[:16]}")
            else:
                st.caption("No actions logged yet.")

# ==========================================
# HEADER
# ==========================================
st.markdown(
    "<h1 class='main-header'>🛡️ GST Reconciliation Tool "
    "<span style='font-size:20px;color:#0f52ba;vertical-align:middle;"
    "background:#e3f2fd;padding:4px 10px;border-radius:15px;'>Enterprise v4.0</span></h1>",
    unsafe_allow_html=True
)
st.markdown(
    "<p class='sub-header'>Automated matching, compliance checking, and discrepancy reporting — B2B, B2BA & CDNR.</p>",
    unsafe_allow_html=True
)

# ── License status banner ─────────────────────────────────────────────────────
if "lic_banner" in st.session_state:
    _mode, _msg = st.session_state["lic_banner"]
    if _mode == "trial":
        st.warning(f"⏳ **Trial Mode** — {_msg}  |  Enter an activation key to unlock full access.")
        with st.expander("🔑 Enter Activation Key", expanded=False):
            _k = st.text_input("Activation Key", placeholder="XXXX-XXXX-XXXX-XXXX", key="inline_key")
            if st.button("Activate Now", key="inline_activate"):
                if _k:
                    from modules.license_manager import activate_key as _activate
                    _r = _activate(_k)
                    if _r["success"]:
                        st.success(_r["message"])
                        st.session_state.pop("lic_banner", None)
                        st.rerun()
                    else:
                        st.error(_r["message"])
    elif _mode == "active":
        st.success(f"✅ **Licensed** — {_msg}")

# ==========================================
# GSTR-2B MULTI-FILE MERGER (TOP CORNER)
# ==========================================

import copy as _copy
import openpyxl as _openpyxl

# NIC GSTR-2B: row number where actual data starts (everything before = header template)
_GSTR2B_DATA_START = {
    'B2B': 7, 'B2BA': 8, 'B2B-CDNR': 7, 'B2B-CDNRA': 8,
    'IMPG': 7, 'IMPGSEZ': 7, 'ISD': 7, 'ISDA': 8,
    'ECOMM': 7, 'ECOMMA': 8,
    'B2B(REJECTED)': 7, 'B2BA(REJECTED)': 8,
    'B2B-CDNR(REJECTED)': 7, 'B2B-CDNRA(REJECTED)': 8,
    'ECO(REJECTED)': 7, 'ECOA(REJECTED)': 8, 'ISD(REJECTED)': 7,
}

def _copy_ws_row(src_ws, dst_ws, src_row, dst_row):
    """Copy one row with cell values + styles from src to dst worksheet."""
    for col_idx, src_cell in enumerate(src_ws[src_row], 1):
        dst_cell = dst_ws.cell(row=dst_row, column=col_idx)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font      = _copy.copy(src_cell.font)
            dst_cell.fill      = _copy.copy(src_cell.fill)
            dst_cell.border    = _copy.copy(src_cell.border)
            dst_cell.alignment = _copy.copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

def _copy_merged_cells(src_ws, dst_ws, up_to_row):
    """Copy merged cell regions that lie within header rows."""
    for rng in src_ws.merged_cells.ranges:
        if rng.min_row <= up_to_row:
            try:
                dst_ws.merge_cells(str(rng))
            except Exception:
                pass

def _get_sheet(wb, norm_name):
    """Find a sheet in workbook by normalised (upper) name."""
    for s in wb.sheetnames:
        if s.strip().upper() == norm_name:
            return s
    return None

def _extract_data_rows(wb, norm_name, data_start):
    """Return list-of-tuples for data rows in the given sheet."""
    sheet = _get_sheet(wb, norm_name)
    if sheet is None:
        return []
    ws = wb[sheet]
    rows = []
    for r in range(data_start, ws.max_row + 1):
        row = tuple(ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1))
        if any(v is not None for v in row):
            rows.append(row)
    return rows

def merge_gstr2b_files(uploaded_files):
    """
    Properly merges multiple NIC-format GSTR-2B Excel files.
    - Preserves rows 1–(data_start-1) from the first file as template (title, sub-headers, merged cells)
    - Stacks data rows from all files, deduplicating by GSTIN + Invoice No + Date
    Returns (bytes, error_message).
    """
    workbooks = []
    for f in uploaded_files:
        try:
            f.seek(0)
            workbooks.append(_openpyxl.load_workbook(f, data_only=True))
        except Exception as e:
            st.warning(f"⚠️ Could not open `{f.name}`: {e}")

    if not workbooks:
        return None, "No readable GSTR-2B files."

    template_wb = workbooks[0]
    out_wb = _openpyxl.Workbook()
    out_wb.remove(out_wb.active)

    # Collect sheet order from first file, keep same sequence
    for sheet_name in template_wb.sheetnames:
        norm = sheet_name.strip().upper()
        data_start = _GSTR2B_DATA_START.get(norm)
        src_ws = template_wb[sheet_name]
        dst_ws = out_wb.create_sheet(title=sheet_name)

        # Copy column widths
        for col_ltr, dim in src_ws.column_dimensions.items():
            dst_ws.column_dimensions[col_ltr].width = dim.width or 14
        for row_idx, dim in src_ws.row_dimensions.items():
            if dim.height:
                dst_ws.row_dimensions[row_idx].height = dim.height

        if data_start is None:
            # Non-data sheet (Read Me, ITC Available…) — copy as-is from template
            for r in range(1, src_ws.max_row + 1):
                _copy_ws_row(src_ws, dst_ws, r, r)
            _copy_merged_cells(src_ws, dst_ws, src_ws.max_row)
        else:
            # 1 — Copy header rows exactly from template
            for r in range(1, data_start):
                _copy_ws_row(src_ws, dst_ws, r, r)
            _copy_merged_cells(src_ws, dst_ws, data_start - 1)

            # 2 — Collect + deduplicate data from all workbooks
            seen_keys = set()
            all_data  = []
            for wb in workbooks:
                for row in _extract_data_rows(wb, norm, data_start):
                    # Dedup key: col0=GSTIN, col2=Invoice/Note No, col4=Date
                    try:
                        key = (str(row[0]).strip().upper(),
                               str(row[2]).strip().upper(),
                               str(row[4]).strip())
                    except Exception:
                        key = str(row[:5])
                    if key not in seen_keys:
                        seen_keys.add(key)
                        all_data.append(row)

            # 3 — Write merged data
            for offset, data_row in enumerate(all_data):
                write_row = data_start + offset
                for col_idx, val in enumerate(data_row, 1):
                    dst_ws.cell(row=write_row, column=col_idx).value = val

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), None


# ── MERGER UI — top right corner via columns ─────────────────────────────────
_merger_col, _merger_btn_col = st.columns([5, 1])
with _merger_btn_col:
    if st.button("🔀 Merge 2B Files", help="Merge multiple GSTR-2B files into one", use_container_width=True):
        st.session_state['show_merger'] = not st.session_state.get('show_merger', False)

if st.session_state.get('show_merger', False):
    with st.container():
        st.markdown("""
        <div style='background: linear-gradient(135deg, #1a237e, #1565c0); color:white;
                    padding:16px 24px; border-radius:12px; margin-bottom:16px;'>
            <h3 style='margin:0;color:white;'>🔀 GSTR-2B Multi-File Merger</h3>
            <p style='margin:4px 0 0 0; font-size:13px; opacity:0.85;'>
                Combine multiple months' GSTR-2B files into a single consolidated Excel.
                Automatically deduplicates by GSTIN + Invoice No + Taxable Value.
            </p>
        </div>
        """, unsafe_allow_html=True)

        merger_files = st.file_uploader(
            "Upload 2 or more GSTR-2B Excel files",
            type=['xlsx'],
            accept_multiple_files=True,
            key="gstr2b_merger_uploader",
            help="Upload GSTR-2B files downloaded from GST Portal (standard NIC format)"
        )

        if merger_files and len(merger_files) >= 2:
            m1, m2, m3 = st.columns(3)
            m1.metric("Files Selected", len(merger_files))
            total_size = sum(f.size for f in merger_files) / 1024
            m2.metric("Total Size", f"{total_size:.1f} KB")
            m3.metric("Format", "GSTR-2B NIC Excel")

            st.markdown("**Files to merge:**")
            for f in merger_files:
                st.caption(f"📄 {f.name}  ({f.size/1024:.1f} KB)")

            if st.button("▶️ Run Merge & Download", type="primary", use_container_width=True, key="run_merger"):
                with st.spinner("Merging files... deduplicating invoices..."):
                    merged_bytes, err = merge_gstr2b_files(merger_files)
                if err:
                    st.error(f"❌ Merge failed: {err}")
                elif merged_bytes:
                    st.success(f"✅ Merged {len(merger_files)} files successfully! Duplicates removed automatically.")
                    st.download_button(
                        label="📥 Download Merged GSTR-2B",
                        data=merged_bytes,
                        file_name=f"GSTR2B_Merged_{len(merger_files)}files.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
        elif merger_files and len(merger_files) == 1:
            st.info("☝️ Please upload at least 2 GSTR-2B files to merge.")

        st.markdown("---")

# ==========================================
# STAGE 1 — SETUP
# ==========================================
if st.session_state.app_stage == 'setup':

    # Template download
    with st.expander("📥 Download Purchase Register Template (Standard Format)", expanded=False):
        st.info("Use this template to prepare your Purchase Register in the correct format before uploading.")
        st.download_button(
            "📥 Download Template Excel",
            data=make_template_excel(),
            file_name="Purchase_Register_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with st.container():
        st.markdown("### 📂 Data Import")

        # ── Software Profile Selector ─────────────────────────────────────────
        st.markdown("#### 🧩 Step 0 — Select Your Accounting Software")
        st.caption("Tell us which software your Purchase Register came from — we'll auto-map the columns for you.")
        software_names = list(SOFTWARE_COLUMN_PROFILES.keys())
        selected_software = st.selectbox(
            "Purchase Register Source Software",
            software_names,
            index=0,
            help="This pre-fills column mapping automatically based on your software's export format.",
            key="software_profile"
        )
        software_profile = SOFTWARE_COLUMN_PROFILES[selected_software]

        # Software badges
        badge_colors = {
            "Standard / GST Portal": "#1565c0",
            "Tally Prime / ERP 9": "#2e7d32",
            "Zoho Books": "#e65100",
            "BUSY Accounting": "#6a1b9a",
            "EasyGST / ClearTax": "#00838f",
            "Marg ERP": "#c62828",
            "Custom / Local Software": "#37474f",
        }
        badge_color = badge_colors.get(selected_software, "#37474f")
        st.markdown(
            f"<span style='background:{badge_color};color:white;padding:4px 12px;"
            f"border-radius:20px;font-size:13px;font-weight:600;'>✓ Profile: {selected_software}</span>",
            unsafe_allow_html=True
        )
        st.markdown("")

        col1, col2 = st.columns(2)
        with col1:
            st.info(f"**Step 1: Books Data** *(from {selected_software})*")
            file_books = st.file_uploader("Upload Purchase Register (Excel/CSV)", type=['xlsx', 'csv'], key="b_up")
        with col2:
            st.info("**Step 2: GSTR-2B Data**")
            file_gst   = st.file_uploader("Upload GSTR-2B (Excel/CSV)", type=['xlsx', 'csv'], key="g_up")

    if file_books and file_gst:

        # Save bytes for CDNR tab
        st.session_state['file_books_bytes'] = file_books.read(); file_books.seek(0)
        st.session_state['file_gst_bytes']   = file_gst.read();   file_gst.seek(0)

        st.divider()
        final_books_map = {}
        final_gst_map   = {}

        # Load B2B data
        df_b_raw = load_data_preview(file_books)
        df_g_raw = load_data_preview(file_gst)

        # --- DATA CONFIDENCE PANEL ---
        if df_b_raw is not None and df_g_raw is not None:
            st.markdown("### 📊 Data Confidence Check")
            st.caption("Verify the numbers below match your source files before running reconciliation.")
            b_summary = make_data_summary(df_b_raw, "Purchase Register (Books)")
            g_summary = make_data_summary(df_g_raw, "GSTR-2B Portal")
            conf_c1, conf_c2 = st.columns(2)
            with conf_c1:
                st.markdown(f"""
                <div class="confidence-box">
                  <div class="confidence-title">📚 {b_summary['label']}</div>
                  Rows loaded: <b>{b_summary['rows']}</b> &nbsp;|&nbsp;
                  Unique Invoices: <b>{b_summary['invoices']}</b> &nbsp;|&nbsp;
                  Unique GSTINs: <b>{b_summary['gstins']}</b><br>
                  Total Taxable Value: <b>₹ {b_summary['taxable']:,.2f}</b>
                </div>""", unsafe_allow_html=True)
            with conf_c2:
                st.markdown(f"""
                <div class="confidence-box" style="border-color:#1565c0;background:#e3f2fd;">
                  <div class="confidence-title" style="color:#0d47a1;">🏛️ {g_summary['label']}</div>
                  Rows loaded: <b>{g_summary['rows']}</b> &nbsp;|&nbsp;
                  Unique Invoices: <b>{g_summary['invoices']}</b> &nbsp;|&nbsp;
                  Unique GSTINs: <b>{g_summary['gstins']}</b><br>
                  Total Taxable Value: <b>₹ {g_summary['taxable']:,.2f}</b>
                </div>""", unsafe_allow_html=True)
            st.session_state['data_summary_books'] = b_summary
            st.session_state['data_summary_gst']   = g_summary
            st.divider()

        # B2BA amendments
        file_gst.seek(0)
        df_b2ba, status_msg = smart_read_b2ba(file_gst)
        if df_b2ba is not None and not df_b2ba.empty:
            st.info(f"⚡ Processing B2B Amendments... Found {len(df_b2ba)} entries in B2BA.")
            df_g_raw, deleted_count, added_count = process_amendments(df_g_raw, df_b2ba)
            st.success(f"✅ B2B Amendments Applied: Removed {deleted_count} old invoices, Added {added_count} revised.")
        elif status_msg and "Critical" in str(status_msg):
            st.warning(status_msg)

        file_gst.seek(0)
        try:
            _xls_check = pd.ExcelFile(file_gst)
            _has_cdnr  = any('cdnr' in s.lower() for s in _xls_check.sheet_names)
        except:
            _has_cdnr  = False
        file_gst.seek(0)
        if _has_cdnr:
            st.info("📋 CDNR sheet detected in GSTR-2B. Run **CDNR Reconciliation** from **Tab 2** after B2B recon.")

        # Auto-detect metadata
        det_fy, det_period, det_gstin, det_name = "2025 - 2026", "March", "24ABLPL3808P1Z8", "SHYAM CREATION"
        meta_fy, meta_period, meta_gstin, meta_name = extract_meta_from_readme(file_gst)
        if meta_gstin: det_gstin  = meta_gstin
        if meta_name:  det_name   = meta_name
        if meta_fy:    det_fy     = meta_fy
        if meta_period: det_period = meta_period

        # Column Mapper — software-profile-aware
        with st.expander("🛠️ Column Mapping Configuration", expanded=False):
            st.caption(f"Auto-mapped using **{selected_software}** profile. Verify and adjust if needed.")
            cols_books            = list(df_b_raw.columns) if df_b_raw is not None else []
            cols_gst              = list(df_g_raw.columns) if df_g_raw is not None else []
            cols_books_with_blank = ["<No Column / Blank>"] + cols_books

            def smart_find_with_profile(field, available_cols, profile_aliases, fallback_map):
                """Find best column using software-profile aliases first, then generic fallback."""
                profile_candidates = profile_aliases.get(field, [])
                available_lower = [str(c).lower() for c in available_cols]
                # Try exact match from profile aliases
                for candidate in profile_candidates:
                    if candidate in available_cols:
                        return candidate
                    # Try case-insensitive
                    try:
                        idx = available_lower.index(candidate.lower())
                        return available_cols[idx]
                    except ValueError:
                        pass
                    # Try partial match
                    for i, col_l in enumerate(available_lower):
                        if candidate.lower() in col_l or col_l in candidate.lower():
                            return available_cols[i]
                # Fall back to generic matcher
                return find_best_match(field, available_cols, fallback_map)

            map_col1, map_col2 = st.columns(2)
            with map_col1:
                st.markdown(f"#### 📚 Purchase Register ({selected_software})")
                for field in REQUIRED_FIELDS:
                    suggested = smart_find_with_profile(field, cols_books_with_blank, software_profile, FIXED_BOOKS_MAPPING)
                    idx = cols_books_with_blank.index(suggested) if suggested in cols_books_with_blank else 0
                    # Show confidence indicator
                    is_profile_match = suggested != "<No Column / Blank>" and suggested in cols_books_with_blank
                    label = f"{field} {'✅' if is_profile_match else '⚠️'}"
                    final_books_map[field] = st.selectbox(label, cols_books_with_blank, index=idx, key=f"b_{field}")
            with map_col2:
                st.markdown("#### 🏛️ GSTR-2B (Portal)")
                for field in REQUIRED_FIELDS:
                    suggested = find_best_match(field, cols_gst, FIXED_GST_MAPPING)
                    idx = cols_gst.index(suggested) if suggested in cols_gst else 0
                    final_gst_map[field] = st.selectbox(f"{field} (GST)", cols_gst, index=idx, key=f"g_{field}")

            # Mapping confidence score
            mapped_count = sum(1 for v in final_books_map.values() if v != "<No Column / Blank>")
            total_fields = len(REQUIRED_FIELDS)
            confidence_pct = int(mapped_count / total_fields * 100)
            conf_color = "#4caf50" if confidence_pct >= 80 else "#ff9800" if confidence_pct >= 50 else "#f44336"
            st.markdown(
                f"<div style='margin-top:10px;padding:8px 16px;background:#f5f5f5;border-radius:8px;'>"
                f"Mapping confidence: <strong style='color:{conf_color}'>{confidence_pct}% ({mapped_count}/{total_fields} fields mapped)</strong>"
                f"</div>",
                unsafe_allow_html=True
            )

        st.divider()
        st.markdown("### ⚙️ Reconciliation Settings")

        c1, c2, c3, c4 = st.columns(4)
        with c1: gstin_input  = st.text_input("GSTIN", det_gstin)
        with c2: name_input   = st.text_input("Client Name", det_name)
        with c3: fy_input     = st.text_input("Financial Year", det_fy)
        with c4: period_input = st.text_input("Period", det_period)

        # GSTIN Validation
        if gstin_input and not validate_gstin(gstin_input):
            st.warning(f"⚠️ GSTIN format looks invalid: `{gstin_input}`. Expected format: 22AAAAA0000A1Z5")

        t1, t2, t3 = st.columns([1, 1, 2])
        with t1: tolerance_input = st.number_input("Global Tolerance (₹)", value=5.0, step=1.0, help="Default allowable difference for matching")
        with t2: smart_mode_input = st.checkbox("Enable Smart Suggestions (Fuzzy Logic)", value=False)

        # Per-vendor tolerance
        with st.expander("⚙️ Per-Vendor Tolerance Overrides (Advanced)", expanded=False):
            st.caption("Override the global tolerance for specific suppliers (GSTIN). Leave empty to use global.")
            vendor_tol_data = st.data_editor(
                pd.DataFrame(list(st.session_state.vendor_tolerances.items()) or [('', 5.0)],
                             columns=['GSTIN', 'Tolerance (₹)']),
                num_rows="dynamic", use_container_width=True, key="vendor_tol_editor"
            )
            if st.button("💾 Save Tolerances"):
                new_tol = {}
                for _, row in vendor_tol_data.iterrows():
                    if row['GSTIN'] and str(row['GSTIN']).strip():
                        new_tol[str(row['GSTIN']).strip().upper()] = float(row['Tolerance (₹)'])
                st.session_state.vendor_tolerances = new_tol
                st.success(f"✅ Saved {len(new_tol)} tolerance overrides.")

        if st.button("🚀 Run Reconciliation Engine", type="primary", use_container_width=True):

            # Validate GSTINs in data
            if df_b_raw is not None:
                gstin_col_b = next((c for c in df_b_raw.columns if 'gstin' in c.lower()), None)
                if gstin_col_b:
                    invalid_gstins = df_b_raw[~df_b_raw[gstin_col_b].astype(str).apply(validate_gstin)][gstin_col_b].unique()
                    if len(invalid_gstins) > 0:
                        st.warning(f"⚠️ {len(invalid_gstins)} invalid GSTIN(s) found in Books data. They will be processed but may not match correctly.")

            books_rename_map = {v: k for k, v in final_books_map.items() if v != "<No Column / Blank>"}
            gst_rename_map   = {v: k for k, v in final_gst_map.items()}

            df_b_clean = df_b_raw.rename(columns=books_rename_map)
            df_g_clean = df_g_raw.rename(columns=gst_rename_map)

            df_b_clean = standardize_invoice_numbers(df_b_clean, "Invoice Number")
            df_g_clean = standardize_invoice_numbers(df_g_clean, "Invoice Number")

            for req_field, mapped_val in final_books_map.items():
                if mapped_val == "<No Column / Blank>":
                    df_b_clean[req_field] = np.nan

            df_b_clean = df_b_clean[[k for k in REQUIRED_FIELDS.keys() if k in df_b_clean.columns]]
            df_g_clean = df_g_clean[[k for k in REQUIRED_FIELDS.keys() if k in df_g_clean.columns]]
            df_b_clean = df_b_clean.loc[:, ~df_b_clean.columns.duplicated()]
            df_g_clean = df_g_clean.loc[:, ~df_g_clean.columns.duplicated()]

            st.session_state['df_b_clean']  = df_b_clean
            st.session_state['df_g_clean']  = df_g_clean
            st.session_state['tolerance']   = tolerance_input
            st.session_state['smart_mode']  = smart_mode_input
            st.session_state['meta_gstin']  = gstin_input
            st.session_state['meta_name']   = name_input
            st.session_state['meta_fy']     = fy_input
            st.session_state['meta_period'] = period_input
            st.session_state.cdnr_result    = None
            st.session_state.cdnr_summary   = None
            st.session_state.current_recon_id = None
            st.session_state.app_stage = 'processing'
            st.rerun()

# ==========================================
# STAGE 2 — PROCESSING
# ==========================================
elif st.session_state.app_stage == 'processing':
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("<h3 style='text-align:center;color:#444;'>🤖 The Reconciliation Engine is processing your data...</h3>", unsafe_allow_html=True)
    show_processing_animation()

    df_b  = st.session_state['df_b_clean']
    df_g  = st.session_state['df_g_clean']
    tol   = st.session_state['tolerance']
    smart = st.session_state['smart_mode']

    time.sleep(0.5)
    result, df_b_rem, df_g_rem = run_reconciliation(df_b, df_g, tol, st.session_state.manual_matches, smart)
    result['Final_Taxable'] = result['Taxable Value_BOOKS'].fillna(result['Taxable Value_GST']).fillna(0)

    meta = {
        'gstin':  st.session_state['meta_gstin'],
        'name':   st.session_state['meta_name'],
        'fy':     st.session_state['meta_fy'],
        'period': st.session_state['meta_period']
    }
    recon_id = save_reconciliation(meta, result)
    st.session_state.current_recon_id   = recon_id
    st.session_state.current_client_path = get_client_path(meta['name'], meta['gstin'], meta['fy'], meta['period'])
    st.session_state['last_result'] = result
    log_action(recon_id, 'new_recon', {'invoices': len(result), 'tolerance': tol})
    st.session_state.app_stage = 'results'
    st.rerun()

# ==========================================
# STAGE 3 — RESULTS
# ==========================================
elif st.session_state.app_stage == 'results':

    c_res1, c_res2 = st.columns([3, 1])
    with c_res1:
        st.success(f"✅ B2B Reconciliation Complete for **{st.session_state['meta_name']}** ({st.session_state['meta_period']})")
    with c_res2:
        col_btn1, col_btn2 = st.columns(2)
        with col_btn1:
            if st.button("📁 Folder"):
                if st.session_state.current_client_path:
                    import platform as _plat
                    if _plat.system() == "Linux":
                        st.info("📁 Files are saved in session memory on cloud. Use the download buttons above to save reports.")
                    else:
                        open_folder(st.session_state.current_client_path)
        with col_btn2:
            if st.button("🔄 New"):
                st.session_state.app_stage = 'setup'
                st.rerun()

    result = st.session_state['last_result']
    df_b   = st.session_state['df_b_clean']
    df_g   = st.session_state['df_g_clean']
    gstin  = st.session_state['meta_gstin']
    name   = st.session_state['meta_name']
    fy     = st.session_state['meta_fy']
    period = st.session_state['meta_period']

    # Safe display copy
    result_display = result.copy()
    if 'Invoice Date_BOOKS' in result_display.columns:
        result_display['Invoice Date_BOOKS'] = pd.to_datetime(
            result_display['Invoice Date_BOOKS'], dayfirst=True, errors='coerce'
        ).dt.strftime('%d/%m/%Y').fillna(result_display['Invoice Date_BOOKS'])
    if 'Invoice Date_GST' in result_display.columns:
        result_display['Invoice Date_GST'] = pd.to_datetime(
            result_display['Invoice Date_GST'], dayfirst=True, errors='coerce'
        ).dt.strftime('%d/%m/%Y').fillna(result_display['Invoice Date_GST'])

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 Dashboard & Scorecard",
        "📋 CDNR Matching",
        "📝 Detailed Data",
        "🏢 Supplier Wise",
        "🔗 Manual Matcher",
        "💬 Vendor Comms",
    ])

    # ─────────────────────────────────────────────────────
    # TAB 1 — DASHBOARD
    # ─────────────────────────────────────────────────────
    with tab1:
        total_books_val = df_b['Taxable Value'].sum() if 'Taxable Value' in df_b.columns else 0
        total_gst_val   = df_g['Taxable Value'].sum() if 'Taxable Value' in df_g.columns else 0
        diff_val        = total_books_val - total_gst_val

        m1, m2, m3 = st.columns(3)
        m1.metric("Books Taxable Value (B2B)",   f"₹ {total_books_val:,.2f}", delta="Input Credit Available")
        m2.metric("GSTR-2B Taxable Value (B2B)", f"₹ {total_gst_val:,.2f}",  delta="Portal Data", delta_color="off")
        m3.metric("Difference",                  f"₹ {diff_val:,.2f}",       delta="Gap to Reconcile", delta_color="inverse")

        # --- Data Confidence Panel (after recon) ---
        b_sum = st.session_state.get('data_summary_books')
        g_sum = st.session_state.get('data_summary_gst')
        if b_sum or g_sum:
            with st.expander("📋 Data Loaded Summary — Confidence Check", expanded=False):
                dc1, dc2 = st.columns(2)
                if b_sum:
                    dc1.markdown(f"""
                    **📚 Purchase Register (Books)**  
                    Rows: `{b_sum['rows']}` | Invoices: `{b_sum['invoices']}` | GSTINs: `{b_sum['gstins']}`  
                    Total Taxable: **₹ {b_sum['taxable']:,.2f}**
                    """)
                if g_sum:
                    dc2.markdown(f"""
                    **🏛️ GSTR-2B Portal**  
                    Rows: `{g_sum['rows']}` | Invoices: `{g_sum['invoices']}` | GSTINs: `{g_sum['gstins']}`  
                    Total Taxable: **₹ {g_sum['taxable']:,.2f}**
                    """)

        # --- Combined ITC Banner (B2B + CDNR) ---
        if st.session_state.cdnr_result is not None and st.session_state.cdnr_summary:
            cdnr_itc    = st.session_state.cdnr_summary.get('net_itc_impact', 0)
            net_itc     = total_books_val + cdnr_itc
            st.info(
                f"💡 **Net ITC Summary** | B2B Gross ITC: ₹ {total_books_val:,.2f} "
                f"| CDNR Adjustments: ₹ {cdnr_itc:,.2f} "
                f"| **Net Eligible ITC: ₹ {net_itc:,.2f}**"
            )

        st.divider()

        col_chart, col_table = st.columns([1, 2])
        summary_df = result['Recon_Status'].value_counts().reset_index()
        summary_df.columns = ['Status', 'Count']

        with col_chart:
            st.markdown("**Status Distribution**")
            chart = alt.Chart(summary_df).mark_arc(innerRadius=60).encode(
                theta=alt.Theta("Count", stack=True),
                color=alt.Color("Status", scale=alt.Scale(scheme='category20'), legend=None),
                tooltip=["Status", "Count"]
            )
            st.altair_chart(chart, use_container_width=True)

        with col_table:
            st.markdown("**Financial Impact Analysis**")
            val_summary   = result.groupby('Recon_Status')['Final_Taxable'].sum().reset_index()
            final_summary = pd.merge(summary_df, val_summary, left_on='Status', right_on='Recon_Status')[
                ['Status', 'Count', 'Final_Taxable']]
            st.dataframe(
                final_summary, use_container_width=True, hide_index=True,
                column_config={
                    "Status":        st.column_config.TextColumn("Status", width="large"),
                    "Count":         st.column_config.NumberColumn("Invoices", format="%d"),
                    "Final_Taxable": st.column_config.NumberColumn("Total Taxable", format="₹ %.2f")
                }
            )

        st.markdown("#### 🚨 Risk Radar: Top Non-Compliant Vendors")
        not_in_2b = result[result['Recon_Status'] == 'Invoices Not in GSTR-2B']
        if not not_in_2b.empty:
            risk_df = not_in_2b.groupby('Name of Party').agg(
                Missing_Count=('GSTIN', 'count'),
                Total_Value=('Final_Taxable', 'sum')
            ).reset_index().sort_values('Total_Value', ascending=False).head(5)
            st.dataframe(
                risk_df, hide_index=True, use_container_width=True,
                column_config={
                    "Name of Party": "Vendor Name",
                    "Missing_Count": "Missing Invoices",
                    "Total_Value":   st.column_config.ProgressColumn(
                        "Risk Exposure (₹)", format="₹ %.2f",
                        min_value=0, max_value=float(risk_df['Total_Value'].max()))
                }
            )
        else:
            st.success("Excellent! No vendors found with missing B2B invoices.")

    # ─────────────────────────────────────────────────────
    # TAB 3 — DETAILED DATA
    # ─────────────────────────────────────────────────────
    with tab3:
        filters = ["All Data", "Matched", "Mismatch (Value)", "AI Matched",
                   "Suggestions", "Manually Linked", "Not in 2B", "Not in Books"]
        status_filter = st.selectbox("Filter Data View:", filters, index=0)

        df_view = result_display.copy()

        # BUG FIX: Moved GST Status column logic into the correct branch
        if status_filter == "All Data":
            pass
        elif status_filter == "Matched":
            df_view = result_display[result_display['Recon_Status'].str.contains('Matched', na=False) &
                                     ~result_display['Recon_Status'].str.contains('AI', na=False)]
        elif status_filter == "Mismatch (Value)":
            df_view = result_display[result_display['Recon_Status'].str.contains('Mismatch', na=False)]
        elif status_filter == "AI Matched":
            df_view = result_display[result_display['Recon_Status'].str.contains('AI', na=False)]
        elif status_filter == "Suggestions":
            df_view = result_display[result_display['Recon_Status'].str.contains('Suggestion', na=False)].copy()
            # Show GSTIN match status for suggestions
            if 'GSTIN_BOOKS' in df_view.columns and 'GSTIN_GST' in df_view.columns:
                df_view.insert(0, 'GSTIN Match?',
                               np.where(df_view['GSTIN_BOOKS'] == df_view['GSTIN_GST'], '✅ Same', '❌ Different'))
        elif status_filter == "Manually Linked":
            df_view = result_display[result_display['Recon_Status'].str.contains('Manual', na=False)]
        elif status_filter == "Not in 2B":
            df_view = result_display[result_display['Recon_Status'] == "Invoices Not in GSTR-2B"]
        elif status_filter == "Not in Books":
            df_view = result_display[result_display['Recon_Status'] == "Invoices Not in Purchase Books"]

        st.dataframe(
            df_view, use_container_width=True,
            column_config={
                "Taxable Value_BOOKS": st.column_config.NumberColumn(format="₹ %.2f"),
                "Taxable Value_GST":   st.column_config.NumberColumn(format="₹ %.2f"),
                "Final_Taxable":       st.column_config.NumberColumn(format="₹ %.2f"),
            }
        )

    # ─────────────────────────────────────────────────────
    # TAB 4 — SUPPLIER WISE
    # ─────────────────────────────────────────────────────
    with tab4:
        pivot = result.groupby('Name of Party').agg(
            Total_Invoices  =('GSTIN_BOOKS' if 'GSTIN_BOOKS' in result.columns else 'GSTIN', 'count'),
            Taxable_Value   =('Final_Taxable', 'sum'),
            Unmatched_Count =('Recon_Status', lambda x: x.str.contains('Not in', na=False).sum())
        ).reset_index().sort_values('Unmatched_Count', ascending=False)
        st.dataframe(
            pivot, use_container_width=True, hide_index=True,
            column_config={
                "Name of Party":  "Vendor",
                "Total_Invoices": st.column_config.NumberColumn("Total Inv"),
                "Taxable_Value":  st.column_config.NumberColumn("Total Business (₹)", format="₹ %.2f"),
                "Unmatched_Count": st.column_config.NumberColumn("Discrepancies")
            }
        )

    # ─────────────────────────────────────────────────────
    # TAB 5 — MANUAL MATCHER
    # ─────────────────────────────────────────────────────
    with tab5:
        c1, c2 = st.columns([2, 1])
        with c1: st.write("🔗 **Link Unmatched Invoices Manually**")
        with c2:
            if st.button("Clear All Manual Links", type="secondary"):
                st.session_state.manual_matches = []
                st.session_state.app_stage = 'processing'
                st.rerun()

        unmatched_books = result[result['Recon_Status'] == "Invoices Not in GSTR-2B"].copy()
        unmatched_gst   = result[result['Recon_Status'] == "Invoices Not in Purchase Books"].copy()

        unmatched_books['Label'] = unmatched_books.apply(
            lambda x: f"{x['Name of Party']} | Inv: {x.get('Invoice Number_BOOKS','')} | ₹{x.get('Taxable Value_BOOKS',0)}", axis=1)
        unmatched_gst['Label']   = unmatched_gst.apply(
            lambda x: f"{x['Name of Party']} | Inv: {x.get('Invoice Number_GST','')} | ₹{x.get('Taxable Value_GST',0)}", axis=1)

        if 'Unique_ID_BOOKS' in unmatched_books.columns: unmatched_books['ID'] = unmatched_books['Unique_ID_BOOKS']
        elif 'Unique_ID' in unmatched_books.columns:     unmatched_books['ID'] = unmatched_books['Unique_ID']
        if 'Unique_ID_GST' in unmatched_gst.columns:    unmatched_gst['ID']   = unmatched_gst['Unique_ID_GST']
        elif 'Unique_ID' in unmatched_gst.columns:      unmatched_gst['ID']   = unmatched_gst['Unique_ID']

        col_left, col_mid, col_right = st.columns([1, 0.2, 1])
        with col_left:  b_choice = st.selectbox("Select Invoice from Books",    unmatched_books['Label'].tolist(), index=None)
        with col_mid:   st.markdown("<h2 style='text-align:center;color:#aaa;'>🔗</h2>", unsafe_allow_html=True)
        with col_right: g_choice = st.selectbox("Select Invoice from GSTR-2B", unmatched_gst['Label'].tolist(),   index=None)

        if st.button("Link Selected Pair", type="primary", use_container_width=True):
            if b_choice and g_choice:
                b_id = unmatched_books[unmatched_books['Label'] == b_choice]['ID'].iloc[0]
                g_id = unmatched_gst[unmatched_gst['Label']     == g_choice]['ID'].iloc[0]
                st.session_state.manual_matches.append((b_id, g_id))
                if st.session_state.current_recon_id:
                    log_action(st.session_state.current_recon_id, 'manual_link',
                               {'books_id': str(b_id), 'gst_id': str(g_id),
                                'books_label': b_choice[:80], 'gst_label': g_choice[:80]})
                st.success("Linked! Re-running reconciliation...")
                st.session_state.app_stage = 'processing'
                st.rerun()

        if st.session_state.manual_matches:
            st.info(f"{len(st.session_state.manual_matches)} manual link(s) active.")

    # ─────────────────────────────────────────────────────
    # TAB 6 — VENDOR COMMS
    # ─────────────────────────────────────────────────────
    with tab6:
        st.subheader("💬 Vendor Communication Center")

        with st.expander("🌐 Fix 'Unknown' Vendors via Official GST Portal", expanded=False):
            if 'Name of Party' in result.columns:
                unknown_mask          = result['Name of Party'] == 'Unknown'
                unique_unknown_gstins = result[unknown_mask]['GSTIN'].unique() if not result[unknown_mask].empty else []
            else:
                unique_unknown_gstins = []

            if len(unique_unknown_gstins) > 0:
                st.warning(f"Found {len(unique_unknown_gstins)} unique GSTINs with name 'Unknown'.")
                col_ctrl, col_cap = st.columns([1, 2])
                with col_ctrl:
                    target_gstin_sel = st.selectbox("Select GSTIN to Fetch", unique_unknown_gstins)
                    if st.button("1. Connect & Load Captcha"):
                        try:
                            st.session_state.target_gstin = target_gstin_sel
                            # BUG FIX: Wrapped in try/except — gst_scraper.py may not be installed
                            from modules.gst_scraper import GSTPortalScraper
                            if st.session_state.gst_scraper is None:
                                st.session_state.gst_scraper = GSTPortalScraper()
                                st.session_state.gst_scraper.start_driver()
                            with st.spinner("Connecting..."):
                                img_bytes = st.session_state.gst_scraper.load_page_and_get_captcha()
                                if img_bytes:
                                    st.session_state.captcha_img = img_bytes
                                    st.rerun()
                                else:
                                    st.error("Failed to load captcha.")
                        except ModuleNotFoundError:
                            st.error("⚠️ GST Portal Scraper module not installed. "
                                     "Please add `modules/gst_scraper.py` and install selenium to enable this feature.")
                        except Exception as e:
                            st.error(f"Error: {e}")
                with col_cap:
                    if st.session_state.captcha_img:
                        st.image(st.session_state.captcha_img, caption="Enter Captcha", width=200)
                        with st.form("captcha_form"):
                            user_cap = st.text_input("Enter Characters:", key="cap_input")
                            if st.form_submit_button("Fetch Name"):
                                with st.spinner("Fetching..."):
                                    scraper      = st.session_state.gst_scraper
                                    fetched_name = scraper.perform_search(st.session_state.target_gstin, user_cap)
                                    if "ERROR" in fetched_name:
                                        st.error(fetched_name)
                                    else:
                                        st.success(f"✅ FOUND: {fetched_name}")
                                        mask = (st.session_state['last_result']['GSTIN'] == st.session_state.target_gstin)
                                        st.session_state['last_result'].loc[mask, 'Name of Party'] = fetched_name
                                        if st.session_state.current_recon_id:
                                            log_action(st.session_state.current_recon_id, 'name_change',
                                                       {'gstin': st.session_state.target_gstin, 'new_name': fetched_name})
                                        scraper.close()
                                        st.session_state.gst_scraper = None
                                        st.session_state.captcha_img = None
                                        time.sleep(1)
                                        st.rerun()
            else:
                st.success("All vendors identified! No 'Unknown' names found.")

        with st.expander("✎ Correct Vendor Name (Fix 'Unknown' by GSTIN)", expanded=False):
            all_parties   = sorted(result['Name of Party'].dropna().astype(str).unique().tolist())
            default_ix    = all_parties.index('Unknown') if 'Unknown' in all_parties else 0
            target_vendor = st.selectbox("Select Vendor Name to Fix", all_parties, index=default_ix, key="fix_target_name")
            df_renames    = result[result['Name of Party'] == target_vendor][['GSTIN']].drop_duplicates().reset_index(drop=True)
            df_renames['New Name'] = target_vendor
            st.markdown(f"**Update Names for '{target_vendor}':** (Edit the 'New Name' column below)")
            edited_df = st.data_editor(
                df_renames, hide_index=True,
                column_config={
                    "GSTIN":    st.column_config.TextColumn(disabled=True, width="medium"),
                    "New Name": st.column_config.TextColumn(required=True, width="large")
                },
                use_container_width=True, key="rename_editor"
            )
            if st.button("💾 Save Name Changes", type="primary"):
                changes_count = 0
                for idx, row in edited_df.iterrows():
                    gst     = row['GSTIN']
                    new_val = row['New Name']
                    if new_val != target_vendor:
                        mask = ((st.session_state['last_result']['Name of Party'] == target_vendor) &
                                (st.session_state['last_result']['GSTIN'] == gst))
                        st.session_state['last_result'].loc[mask, 'Name of Party'] = new_val
                        if st.session_state.current_recon_id:
                            log_action(st.session_state.current_recon_id, 'name_change',
                                       {'gstin': gst, 'old_name': target_vendor, 'new_name': new_val})
                        changes_count += 1
                if changes_count > 0:
                    st.success(f"✅ Updated {changes_count} GSTIN(s) successfully!")
                    time.sleep(1)
                    st.rerun()
                else:
                    st.info("No changes detected.")

        result       = st.session_state['last_result']
        issue_vendors = get_vendors_with_issues(result)
        not_in_2b_vendors = result[result['Recon_Status'] == 'Invoices Not in GSTR-2B']['Name of Party'].unique().tolist()
        not_in_2b_vendors = [v for v in not_in_2b_vendors if v and str(v) != 'nan']

        if issue_vendors:
            st.markdown("### 📤 Bulk Actions (Auto-Saved on Download)")
            col_sel1, _ = st.columns([3, 1])
            with col_sel1:
                select_2b_only = st.checkbox(f"Select Only 'Not in 2B' ({len(not_in_2b_vendors)} Vendors)", value=False)
            default_selection     = [v for v in not_in_2b_vendors if v in issue_vendors] if select_2b_only else []
            selected_vendors_bulk = st.multiselect("Select Vendors for Report Generation", issue_vendors, default=default_selection)

            if selected_vendors_bulk:
                c_pdf, c_xls = st.columns(2)
                zip_buffer_pdf = io.BytesIO()
                with zipfile.ZipFile(zip_buffer_pdf, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
                    for v in selected_vendors_bulk:
                        pdf_data = create_vendor_pdf(result, v, name, gstin)
                        zip_file.writestr(f"Notice_{v}.pdf", pdf_data.getvalue())
                filtered_df    = result[result['Name of Party'].isin(selected_vendors_bulk)]
                zip_buffer_xls = generate_vendor_split_zip(filtered_df)
                folder         = st.session_state.current_client_path
                with c_pdf:
                    st.download_button("📄 Download PDF Notices", data=zip_buffer_pdf.getvalue(),
                                       file_name="Notices.zip", type="primary", use_container_width=True,
                                       on_click=save_callback, args=(folder, "Notices.zip", zip_buffer_pdf.getvalue()))
                with c_xls:
                    st.download_button("📊 Download Excel Splits", data=zip_buffer_xls.getvalue(),
                                       file_name="Excels.zip", use_container_width=True,
                                       on_click=save_callback, args=(folder, "Excels.zip", zip_buffer_xls.getvalue()))

        st.divider()
        st.markdown("### 📱 Quick Chat / Email")
        c_vendor, c_mode = st.columns([2, 1])
        with c_vendor: selected_vendor = st.selectbox("Select Single Vendor", issue_vendors)
        with c_mode:   comm_mode       = st.radio("Mode", ["📧 Email", "📱 WhatsApp"], horizontal=True)
        if selected_vendor:
            if comm_mode == "📧 Email":
                subject, body = generate_email_draft(result, selected_vendor, name)
                st.text_input("Subject", value=subject)
                st.markdown("**Email Body:** (Click top-right icon to copy)")
                st.code(body, language='markdown')
            else:
                wa_body = generate_whatsapp_message(result, selected_vendor, name)
                st.markdown("**WhatsApp Message:** (Click top-right icon to copy)")
                st.code(wa_body, language='markdown')
                col_ph, col_btn = st.columns([2, 1])
                with col_ph: phone = st.text_input("Vendor Phone (91...)", value="91")
                with col_btn:
                    st.write(""); st.write("")
                    if phone and len(phone) > 10:
                        encoded_text = urllib.parse.quote(wa_body)
                        st.link_button("🚀 Open in WhatsApp", f"https://wa.me/{phone}?text={encoded_text}")

    # ─────────────────────────────────────────────────────
    # TAB 2 — CDNR MATCHING
    # ─────────────────────────────────────────────────────
    with tab2:
        st.markdown("### 📋 CDNR Reconciliation — Credit & Debit Notes")
        st.caption(
            "Matches Credit/Debit Notes from your **Books (CDNR sheet)** against **GSTR-2B CDNR sheet**. "
            "Values handled separately from B2B — no pollution of B2B KPIs."
        )

        has_files = (
            st.session_state.get('file_books_bytes') is not None and
            st.session_state.get('file_gst_bytes')   is not None
        )

        if not has_files:
            if st.session_state.cdnr_result is not None:
                st.info("📂 Showing CDNR results loaded from history. Re-upload files to run fresh CDNR.")
            else:
                st.warning("⚠️ Original files not available. Click **🔄 New** and re-upload to run CDNR reconciliation.")
        else:
            if st.button("▶️ Run CDNR Reconciliation", type="primary", use_container_width=True, key="run_cdnr"):
                with st.spinner("Reading CDNR sheets, applying CDNRA amendments, and matching notes..."):
                    try:
                        file_b_io = io.BytesIO(st.session_state['file_books_bytes'])
                        file_g_io = io.BytesIO(st.session_state['file_gst_bytes'])
                        cdnr_result, cdnr_summary = process_cdnr_reconciliation(
                            file_b_io, file_g_io,
                            tolerance  = st.session_state.get('tolerance',   5.0),
                            smart_mode = st.session_state.get('smart_mode', False)
                        )
                        st.session_state.cdnr_result  = cdnr_result
                        st.session_state.cdnr_summary = cdnr_summary
                        # Save to DB so history loads restore CDNR results
                        if st.session_state.current_recon_id:
                            save_cdnr_to_history(st.session_state.current_recon_id, cdnr_result, cdnr_summary)
                            log_action(st.session_state.current_recon_id, 'cdnr_run',
                                       {'matched': cdnr_summary.get('matched_count', 0),
                                        'not_in_2b': cdnr_summary.get('not_in_2b_count', 0)})
                    except Exception as e:
                        st.error(f"CDNR Engine Error: {e}")

        if st.session_state.cdnr_result is not None:
            cdnr_result  = st.session_state.cdnr_result
            cdnr_summary = st.session_state.cdnr_summary

            if cdnr_result.empty:
                st.warning(
                    "No CDNR data found. Check that:\n"
                    "- Your Books file has a sheet named exactly **'cdnr'** (lowercase ok)\n"
                    "- Your GSTR-2B file has a **'CDNR'** tab (standard NIC format)"
                )
            else:
                if cdnr_summary.get('amendments_deleted', 0) > 0 or cdnr_summary.get('amendments_added', 0) > 0:
                    st.info(
                        f"⚡ CDNRA Applied: Removed **{cdnr_summary['amendments_deleted']}** old notes, "
                        f"Added **{cdnr_summary['amendments_added']}** revised notes."
                    )

                # KPI Cards (ITC implication section removed)
                k1, k2, k3, k4, k5 = st.columns(5)
                k1.metric("📚 Notes in Books",    cdnr_summary.get('total_books', 0))
                k2.metric("🏛️ Notes in GSTR-2B", cdnr_summary.get('total_gst',   0))
                k3.metric("✅ Matched",            cdnr_summary.get('matched_count', 0))
                k4.metric("⚠️ Value Mismatch",     cdnr_summary.get('mismatch_count', 0))
                k5.metric("❌ Unmatched",
                    cdnr_summary.get('not_in_2b_count', 0) + cdnr_summary.get('not_in_books_count', 0))

                k6, k7 = st.columns(2)
                k6.metric("🔶 Tax Error",  cdnr_summary.get('tax_error_count', 0),
                          help="Taxable matches but IGST/CGST/SGST differs")
                k7.metric("🤖 AI Matched", cdnr_summary.get('ai_matched_count', 0),
                          help="Matched via date/taxable fallback steps")

                st.divider()

                # Filter + CDNR Suggestions tab (shows GSTIN match status like B2B)
                cdnr_filter_opts = [
                    "All Data", "CDNR Matched", "CDNR Matched (Tax Error)",
                    "CDNR AI Matched", "CDNR Mismatch",
                    "CDNR Not in GSTR-2B", "CDNR Not in Books",
                    "⚠️ CDNR Suggestions (Review GSTIN Match)",
                ]
                cdnr_filter = st.selectbox("🔍 Filter CDNR View", cdnr_filter_opts, key="cdnr_filter")

                CDNR_FILTER_MAP = {
                    "CDNR Matched"             : r"CDNR Matched$",
                    "CDNR Matched (Tax Error)" : r"Tax Error",
                    "CDNR AI Matched"          : r"AI Matched",
                    "CDNR Mismatch"            : r"Mismatch",
                    "CDNR Not in GSTR-2B"      : r"Not in GSTR-2B",
                    "CDNR Not in Books"        : r"Not in Books",
                    "⚠️ CDNR Suggestions (Review GSTIN Match)": r"Suggestion",
                }
                df_cdnr_view = cdnr_result.copy()
                if cdnr_filter != "All Data":
                    pat = CDNR_FILTER_MAP.get(cdnr_filter, cdnr_filter)
                    df_cdnr_view = cdnr_result[
                        cdnr_result['Recon_Status_CDNR'].str.contains(pat, regex=True, na=False)
                    ].copy()

                # For Suggestions view: show GSTIN match status column
                if "Suggestion" in cdnr_filter:
                    if 'GSTIN_BOOKS' in df_cdnr_view.columns and 'GSTIN_GST' in df_cdnr_view.columns:
                        df_cdnr_view.insert(0, 'GSTIN Match?',
                            np.where(df_cdnr_view['GSTIN_BOOKS'] == df_cdnr_view['GSTIN_GST'],
                                     '✅ Same GSTIN', '❌ Different GSTIN'))
                    if len(df_cdnr_view) > 0:
                        st.warning(
                            f"⚠️ {len(df_cdnr_view)} CDNR Suggestion(s) found. "
                            "These are cross-GSTIN matches — verify the GSTIN Match column before accepting."
                        )

                st.dataframe(
                    df_cdnr_view, use_container_width=True, hide_index=True,
                    column_config={
                        "Taxable Value_BOOKS": st.column_config.NumberColumn("Taxable (Books)", format="₹ %.2f"),
                        "Taxable Value_GST":   st.column_config.NumberColumn("Taxable (2B)",    format="₹ %.2f"),
                        "Diff_Taxable":        st.column_config.NumberColumn("Diff Taxable",    format="₹ %.2f"),
                        "Diff_IGST":           st.column_config.NumberColumn("Diff IGST",       format="₹ %.2f"),
                        "Diff_CGST":           st.column_config.NumberColumn("Diff CGST",       format="₹ %.2f"),
                        "Diff_SGST":           st.column_config.NumberColumn("Diff SGST",       format="₹ %.2f"),
                        "Recon_Status_CDNR":   st.column_config.TextColumn("Status", width="large"),
                    }
                )

                # Download CDNR Report
                try:
                    cdnr_excel_bytes = generate_cdnr_excel(cdnr_result, gstin, name, fy, period,
                                                           b2b_full_df=result)
                except Exception as _err:
                    st.error(f"Report generation error: {_err}")
                    cdnr_excel_bytes = None

                if cdnr_excel_bytes:
                    cdnr_filename = f"CDNR_Reconciliation_{period}.xlsx"
                    folder        = st.session_state.current_client_path
                    st.download_button(
                        label="📥 Download CDNR Reconciliation Report",
                        data=cdnr_excel_bytes,
                        file_name=cdnr_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                        on_click=save_callback,
                        args=(folder, cdnr_filename, cdnr_excel_bytes),
                    )

    # ─────────────────────────────────────────────────────
    # MAIN B2B REPORT DOWNLOAD
    # ─────────────────────────────────────────────────────
    st.divider()
    excel_data      = generate_excel(result, gstin, name, fy, period,
                                     cdnr_df=st.session_state.get('cdnr_result'))
    report_filename = f"B2B_Reconciliation_Report_{period}.xlsx"
    folder          = st.session_state.current_client_path

    st.download_button(
        label="📥 Download Full B2B Reconciliation Report (with Executive Summary)",
        data=excel_data,
        file_name=report_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
        on_click=save_callback,
        args=(folder, report_filename, excel_data)
    )
